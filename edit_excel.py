from openpyxl import Workbook
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import os
from datetime import datetime
import sqlite3
from sqlite3 import Error
from pytz import timezone 
class Create_connection_xl:
    def __init__(self,db_file):
        self.db_file=db_file
    def create_connection(self):
        """ create a database connection to the SQLite database
            specified by the db_file
        :param db_file: database file
        :return: Connection object or None
        """
        conn = None
        try:
            conn = sqlite3.connect(self.db_file)
        except Error as e:
            print(e)
    
        return conn

class HDD_form_xl:
    def __init__(self,conn,form,span_id,prikey):
        self.conn=conn
        self.form=form
        self.span_id=span_id
        self.prikey=prikey
    def query_tasks(self):
        """
        Query all rows in the tasks table
        :param conn: the Connection object
        :return:
        """
        form=self.form
        prikey=self.prikey
        span_id=self.span_id
    
        cur = self.conn.cursor()
        zone_name='Warangal'
        command="SELECT * FROM hdd WHERE span_id = \'"+span_id+"\' AND prikey >="+ str(prikey)
        cur.execute(command)
        
        
        rows = cur.fetchall()
        if len(rows)==0:
            print('Nothing to Show')
        else:
            date_time=self.make_excel(rows)
            return date_time
    def make_excel(self,data):
        self.data=data
        data=self.data
        wb = Workbook()
        cwd=os.getcwd()
        directory='AT_Report_Excel/R08_HDD.xlsx'
        # Path
        path = os.path.join(cwd, directory)
        # workbook object is created
        wb_obj = openpyxl.load_workbook(path)
        # grab the active worksheet
        ws = wb_obj['R08']
        # Data can be assigned directly to cells
        i=0
        # print(data[i])
        
        while(i<len(data) and i<5):
            if i==0:
                ws['E2'].value = 'TCIL'
                ws['E3'].value= '1'
                ws['E4'].value='1'
                ws['E5'].value='GP'
                ws['E6'].value=data[i][18]+"/"+data[i][20]+"/"+data[i][23]
                ws['E7'].value=''
                ws['E8'].value=data[0][21]
                ws['E9'].value=''

                ws['A12'].value='1'
                # ws['A12']=ws['A12'].font
                ws['B12'].value=data[i][24]
                ws['C12'].value=data[i][25]
                ws['D12'].value=data[i][26]
                ws['E12'].value=data[i][27]
                ws['F12'].value=data[i][29]
                ws['G12'].value=data[i][30]
                ws['H12'].value=data[i][31]
                ws['I12'].value=data[i][32]
                ws['J12'].value=str(data[i][33])

                ws['B19'].value=str(data[i][35])
                ws['D19'].value=str(data[i][36])
                ws['H19'].value=str(data[i][37])
                ws['J19'].value=str(data[i][38])

            if i==1:
                ws['A13'].value='2'
                ws['B13'].value=data[i][24]
                ws['C13'].value=data[i][25]
                ws['D13'].value=data[i][26]
                ws['E13'].value=data[i][27]
                ws['F13'].value=data[i][29]
                ws['G13'].value=data[i][30]
                ws['H13'].value=data[i][31]
                ws['I13'].value=data[i][32]
                ws['J13'].value=str(data[i][33])
                ws['B20'].value=str(data[i][35])
                ws['D20'].value=str(data[i][36])
                ws['H20'].value=str(data[i][37])
                ws['J20'].value=str(data[i][38])

            if i==2:
                ws['A14'].value='3'
                ws['B14'].value=data[i][24]
                ws['C14'].value=data[i][25]
                ws['D14'].value=data[i][26]
                ws['E14'].value=data[i][27]
                ws['F14'].value=data[i][29]
                ws['G14'].value=data[i][30]
                ws['H14'].value=data[i][31]
                ws['I14'].value=data[i][32]
                ws['J14'].value=str(data[i][33])
                ws['B21'].value=str(data[i][35])
                ws['D21'].value=str(data[i][36])
                ws['H21'].value=str(data[i][37])
                ws['J21'].value=str(data[i][38])

            if i==3:
                ws['A15'].value='4'
                ws['B15'].value=data[i][24]
                ws['C15'].value=data[i][25]
                ws['D15'].value=data[i][26]
                ws['E15'].value=data[i][27]
                ws['F15'].value=data[i][29]
                ws['G15'].value=data[i][30]
                ws['H15'].value=data[i][31]
                ws['I15'].value=data[i][32]
                ws['J15'].value=str(data[i][33])
                ws['B22'].value=str(data[i][35])
                ws['D22'].value=str(data[i][36])
                ws['H22'].value=str(data[i][37])
                ws['J22'].value=str(data[i][38])

            if i==4:
                ws['A16'].value='5'
                ws['B16'].value=data[i][24]
                ws['C16'].value=data[i][25]
                ws['D16'].value=data[i][26]
                ws['E16'].value=data[i][27]
                ws['F16'].value=data[i][29]
                ws['G16'].value=data[i][30]
                ws['H16'].value=data[i][31]
                ws['I16'].value=data[i][32]
                ws['J16'].value=str(data[i][33])
                ws['B23'].value=str(data[i][35])
                ws['D23'].value=str(data[i][36])
                ws['H23'].value=str(data[i][37])
                ws['J23'].value=str(data[i][38])
            i+=1
        ind_time = datetime.now(timezone("Asia/Kolkata")).strftime('%Y_%m_%d_%H_%M_%S') # this time will be filename
        final_directory="Output\Excel/"+ind_time+self.span_id+".xlsx"
        final_path=os.path.join(cwd, final_directory)    
        wb_obj.save(final_path)   
        return ind_time 
        
# prikey='8'
# span_id='NLG-VAL-5369-M-01-GR04-01'
# form='HDD'
# database = r"data.db"

# create_conn=Create_connection(database)
# conn=create_conn.create_connection()
# hdd_conn=HDD_form(conn,form,span_id,prikey)
# date_time=hdd_conn.query_tasks()



# wb = Workbook()
# cwd=os.getcwd()
# directory='AT_Report_Excel/R8_HDD.xlsx'
# # Path
# path = os.path.join(cwd, directory)
# # print(path)
# # workbook object is created
# wb_obj = openpyxl.load_workbook(path)
# # print(wb_obj.sheetnames)
# # grab the active worksheet
# # sheet1=wb_obj['R08']
# ws = wb_obj['R08']
# # print(sheet1['A2'].value)
# # Data can be assigned directly to cells

# ws['E2'].value = ''



# # Save the file
# wb_obj.save("sample.xlsx")